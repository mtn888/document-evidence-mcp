from __future__ import annotations

from importlib.metadata import version

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from document_evidence_mcp.chunking import chunk_text
from document_evidence_mcp.models import EvidenceDraft, ParseResult
from document_evidence_mcp.parsers.base import ParseContext


class XlsxParser:
    name = "openpyxl"
    version = version("openpyxl")
    extensions = frozenset({".xlsx", ".xlsm", ".xltx", ".xltm"})

    def parse(self, context: ParseContext) -> ParseResult:
        workbook = load_workbook(
            context.source_path,
            read_only=True,
            data_only=False,
            keep_links=False,
        )
        blocks: list[EvidenceDraft] = []
        warnings: list[str] = []
        try:
            for sheet in workbook.worksheets:
                for row_index, row in enumerate(sheet.iter_rows(), start=1):
                    values = [cell.value for cell in row]
                    if not any(value not in (None, "") for value in values):
                        continue
                    cells = [
                        f"{get_column_letter(index)}{row_index}={value}"
                        for index, value in enumerate(values, start=1)
                        if value not in (None, "")
                    ]
                    text = "\t".join(cells)
                    for chunk in chunk_text(
                        text,
                        context.settings.chunk_chars,
                        context.settings.chunk_overlap,
                    ):
                        blocks.append(
                            EvidenceDraft(
                                text=chunk,
                                kind="table_row",
                                section=sheet.title,
                                parser=self.name,
                                metadata={
                                    "coordinate_space": "worksheet_cells",
                                    "sheet": sheet.title,
                                    "row": row_index,
                                },
                            )
                        )
        finally:
            workbook.close()
        return ParseResult(
            parser=self.name,
            parser_version=self.version,
            page_count=len(workbook.sheetnames),
            blocks=blocks,
            warnings=warnings,
        )
